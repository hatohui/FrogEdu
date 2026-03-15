"""
Generate Test_Report_Template.xlsx for FrogEdu project.
Run: pip install openpyxl && python generate_test_report.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette & styles ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MODULE_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
TITLE_FONT  = Font(name="Calibri", bold=True, size=16, color="1F4E79")
BOLD_FONT   = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SEVERITY_FILLS = {
    "Critical": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "High":     PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "Medium":   PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    "Low":      PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
}

# Column widths (A–I)
COL_WIDTHS = [10, 40, 30, 50, 40, 30, 10, 12, 25]
HEADERS = [
    "TC-ID", "Test Case Description", "Pre-conditions",
    "Test Steps", "Expected Outcome", "Actual Outcome",
    "Status", "Severity", "Notes",
]


def style_header_row(ws, row):
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def add_test_case(ws, row, tc):
    for col_idx, val in enumerate(tc, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
    # Colour-code severity column (H = 8)
    sev = tc[7] if len(tc) > 7 else ""
    if sev in SEVERITY_FILLS:
        ws.cell(row=row, column=8).fill = SEVERITY_FILLS[sev]
        ws.cell(row=row, column=8).font = Font(name="Calibri", bold=True, size=11,
                                                color="FFFFFF" if sev == "Critical" else "000000")


def add_module_header(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = MODULE_FONT
    cell.fill = MODULE_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER


def set_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════
#  SHEET 1 – Test Cases
# ═══════════════════════════════════════════════════════════════
ws_tc = wb.active
ws_tc.title = "Test Cases"
set_col_widths(ws_tc)

# Title block
ws_tc.merge_cells("A1:I1")
c = ws_tc.cell(row=1, column=1, value="FrogEdu — Test Report")
c.font = TITLE_FONT
c.alignment = Alignment(horizontal="center", vertical="center")

info_rows = [
    ("Project Name:", "FrogEdu"),
    ("Test Date:", ""),
    ("Tester Name:", ""),
    ("Build/Version:", ""),
    ("Environment:", ""),
]
for idx, (label, val) in enumerate(info_rows, 3):
    ws_tc.cell(row=idx, column=1, value=label).font = BOLD_FONT
    ws_tc.cell(row=idx, column=2, value=val).font = NORMAL_FONT

row = 9  # Start modules here

# ── Module 1: Student Exam Taking Flow ────────────────────────
module1 = [
    ("ST-001", "Student navigates to active exam sessions list",
     "Student is authenticated; ≥1 active exam session exists",
     "1. Log in as Student\n2. Navigate to /app/exam-sessions",
     "Page loads with active/upcoming/ended session tabs. Active sessions show name, exam name, start/end time, retry count.", "", "", "Medium", ""),
    ("ST-002", "Student starts a new exam attempt",
     "Active exam session; student has not exceeded retryTimes",
     "1. Open an active session\n2. Click 'Start Exam'\n3. POST /exam-sessions/{sessionId}/attempts",
     "Attempt created: status=InProgress, attemptNumber increments, exam questions load.", "", "", "High", ""),
    ("ST-003", "Answer a Multiple Choice question (Type=1)",
     "Exam attempt in progress; exam has MC question",
     "1. View MC question\n2. Select one option (A–F)\n3. Verify single selection",
     "Only one option selectable. selectedAnswerIds stores selection. UI highlights choice.", "", "", "High", ""),
    ("ST-004", "Answer a Multiple Answer question (Type=2)",
     "Exam attempt in progress; exam has MA question",
     "1. View MA question\n2. Select 2+ options\n3. Verify multiple checkboxes",
     "Multiple selections allowed. All selected IDs stored in selectedAnswerIds.", "", "", "High", ""),
    ("ST-005", "Answer a True/False question (Type=3)",
     "Exam attempt in progress; exam has TF question",
     "1. View TF question\n2. Select True or False",
     "One of two options (A=True, B=False) selectable. Selection stored.", "", "", "High", ""),
    ("ST-006", "Answer a Fill-in-the-Blank question (Type=5)",
     "Exam attempt in progress; exam has FillInBlank question",
     "1. View FillInBlank question\n2. Type answer",
     "Typed text captured. Server matches against 1–5 accepted answer variants.", "", "", "High", ""),
    ("ST-007", "Answer an Essay question (Type=4)",
     "Exam attempt in progress; exam has Essay question",
     "1. View Essay question\n2. Type free-text in textarea",
     "essayText stores response. No client-side character limit (server validates).", "", "", "High", ""),
    ("ST-008", "Navigate between questions during exam",
     "Exam in progress; 3+ questions",
     "1. Answer Q1\n2. Navigate to Q2\n3. Navigate back to Q1",
     "Answers persisted locally when navigating. Previously answered questions retain selections.", "", "", "Medium", ""),
    ("ST-009", "Submit exam with all questions answered",
     "All questions answered",
     "1. Click 'Submit Exam'\n2. Confirm dialog\n3. POST /exam-sessions/{sessionId}/attempts/submit",
     "Status → Submitted. Score computed for auto-graded types. Essay pending AI grading.", "", "", "Critical", ""),
    ("ST-010", "Submit exam with unanswered questions",
     "≥1 question left unanswered",
     "1. Leave question(s) blank\n2. Click Submit\n3. Confirm",
     "Warning dialog about unanswered questions. On confirm: submitted, blanks score 0.", "", "", "High", ""),
    ("ST-011", "Auto-grading MC/TF/FillInBlank on submission",
     "Exam submitted with auto-gradable answers",
     "1. Submit exam\n2. Check attempt results",
     "score, totalPoints, scorePercentage calculated. isCorrect set per answer.", "", "", "Critical", ""),
    ("ST-012", "AI grading of Essay on submission",
     "Exam submitted with Essay answer",
     "1. Submit exam with essay\n2. AI grades via POST /api/ai/essay/grade",
     "essayFeedback and score returned. Attempt status → Graded once all essays done.", "", "", "Critical", ""),
    ("ST-013", "Exam attempt times out before submission",
     "Student does not submit before endTime",
     "1. Start attempt\n2. Let endTime pass",
     "Status → TimedOut. Partial answers saved. Student cannot continue.", "", "", "High", ""),
    ("ST-014", "Student retries exam within allowed retryTimes",
     "isRetryable=true; attemptNumber < retryTimes",
     "1. Complete first attempt\n2. Return to session\n3. Click Retry",
     "New attempt: attemptNumber+1. Previous answers NOT carried over. Shuffled if configured.", "", "", "High", ""),
    ("ST-015", "Student blocked at max attempts",
     "attemptNumber >= retryTimes",
     "1. Use all attempts\n2. Try to start new attempt",
     "Start/Retry button disabled. Error: max attempts reached. POST returns 400/403.", "", "", "High", ""),
    ("ST-016", "Question shuffling (shouldShuffleQuestions=true)",
     "Session has shuffling enabled",
     "1. Start attempt, note question order\n2. Retry exam",
     "Questions appear in different order across attempts.", "", "", "Low", ""),
    ("ST-017", "Answer option shuffling (shouldShuffleAnswers=true)",
     "Session has answer shuffling enabled; MC question present",
     "1. Start attempt, note option order",
     "MC/MA answer options appear randomized.", "", "", "Low", ""),
]

add_module_header(ws_tc, row, "Module 1: Student Exam Taking Flow")
row += 1
style_header_row(ws_tc, row)
row += 1
for tc in module1:
    add_test_case(ws_tc, row, tc)
    row += 1
row += 1  # blank spacer

# ── Module 2: Exam Generation Flow ────────────────────────────
module2 = [
    ("EG-001", "Teacher creates a new exam with basic info",
     "Teacher is authenticated",
     "1. Navigate to Exams page\n2. Click '+'\n3. Fill Name (3–200), Description (10–1000), Grade, Subject, Topic\n4. Submit",
     "Exam created: isDraft=true. Success toast. Redirect to exam detail/matrix page.", "", "", "Critical", ""),
    ("EG-002", "Cascade loading: Grade → Subject → Topic",
     "Teacher on Create Exam form",
     "1. Select Grade=5 → Subject dropdown loads\n2. Select Subject → Topic dropdown loads",
     "Dropdowns cascade. Placeholders shown when parent empty. Loading spinners visible.", "", "", "High", ""),
    ("EG-003", "Create exam form validation",
     "Teacher on Create Exam form",
     "1. Leave Name blank, blur\n2. Enter 2-char name\n3. Leave Description blank",
     "Validation errors for required fields and min length. Submit disabled until valid.", "", "", "Medium", ""),
    ("EG-004", "Teacher creates matrix for exam",
     "Exam created; teacher chose 'Add Matrix'",
     "1. Open Matrix builder\n2. Add topic rows\n3. Enter quantities per cognitive level\n4. Verify auto-totals\n5. Save (POST /matrices)",
     "Matrix saved with matrixTopics[]. Auto-computed totals and weight %. Success toast.", "", "", "Critical", ""),
    ("EG-005", "Matrix validation — 0 total questions",
     "Matrix builder open; all cells=0",
     "1. Leave all cells 0\n2. Click Save",
     "Rejected: 'Total question count must be > 0'.", "", "", "Medium", ""),
    ("EG-006", "AI generates questions from matrix",
     "Matrix created; AI service running",
     "1. Click 'Generate Questions'\n2. POST /api/ai/generate with matrix_topics",
     "AI returns correct count per topic×level with content, type, level, point, answers[].", "", "", "Critical", ""),
    ("EG-007", "AI generation respects cognitive levels",
     "Matrix specifies per-level quantities",
     "1. Generate questions\n2. Verify cognitive_level per question",
     "Each question's cognitive_level matches requested level.", "", "", "High", ""),
    ("EG-008", "AI generation returns mixed question types",
     "question_type=null in request",
     "1. Generate with null question_type\n2. Check returned types",
     "Mix of MC, MA, TF, FillInBlank, Essay. total_count matches matrix sum.", "", "", "Medium", ""),
    ("EG-009", "AI single question generation",
     "Teacher wants one more question",
     "1. Click 'Generate Single Question'\n2. POST /api/ai/generate-single",
     "Exactly 1 question returned matching requested level and type.", "", "", "High", ""),
    ("EG-010", "Attach matrix to exam",
     "Exam and matrix both exist",
     "1. POST /exams/{examId}/matrix with {matrixId}",
     "Matrix linked. GET /matrices/exam/{examId} returns it.", "", "", "High", ""),
    ("EG-011", "Detach matrix from exam",
     "Exam has attached matrix",
     "1. DELETE /exams/{examId}/matrix",
     "Matrix detached. GET returns 404/empty.", "", "", "Medium", ""),
    ("EG-012", "Teacher manually adds questions to exam",
     "Exam exists; question bank has questions",
     "1. Browse question bank\n2. Select questions\n3. Add to exam",
     "ExamQuestion associations created. Exam detail shows updated list.", "", "", "High", ""),
    ("EG-013", "Publish exam (draft → active)",
     "Exam is draft; has ≥1 question",
     "1. Click 'Publish'\n2. Confirm",
     "isDraft=false, isActive=true. Visible to students when assigned.", "", "", "High", ""),
    ("EG-014", "Publish exam with 0 questions blocked",
     "Exam has no questions",
     "1. Click 'Publish'",
     "Error: 'Cannot publish exam with no questions.' Remains draft.", "", "", "Medium", ""),
]

add_module_header(ws_tc, row, "Module 2: Exam Generation Flow (Matrix-Based)")
row += 1
style_header_row(ws_tc, row)
row += 1
for tc in module2:
    add_test_case(ws_tc, row, tc)
    row += 1
row += 1

# ── Module 3: Exam Session History ────────────────────────────
module3 = [
    ("EH-001", "Student views exam session history page",
     "Student authenticated; ≥1 past attempt",
     "1. Navigate to /app/exam-sessions/history",
     "All participated sessions listed: exam name, date range, attempt count, score.", "", "", "High", ""),
    ("EH-002", "History differentiates session statuses",
     "Student has active, upcoming, ended sessions",
     "1. View session list/tabs\n2. Check badges",
     "Active=green, Upcoming=blue, Ended=gray. Computed from startTime/endTime.", "", "", "Medium", ""),
    ("EH-003", "Student views attempts for a session",
     "Student has 2+ attempts on one session",
     "1. Click session\n2. GET /exam-sessions/{sessionId}/attempts/my",
     "All attempts listed: attemptNumber, dates, score, percentage, status.", "", "", "High", ""),
    ("EH-004", "Attempt statuses display correctly",
     "Attempts in different states",
     "1. View attempts across sessions",
     "InProgress=yellow, Submitted=blue, Graded=green, TimedOut=red.", "", "", "Medium", ""),
    ("EH-005", "Score displayed for graded attempts",
     "Attempt status=Graded",
     "1. View graded attempt",
     "Score as 'score/totalPoints (percentage%)'. E.g. '8/10 (80%)'.", "", "", "High", ""),
    ("EH-006", "Empty state — no history",
     "Student never took exams",
     "1. Navigate to history page",
     "Empty state message displayed. No error thrown.", "", "", "Low", ""),
    ("EH-007", "History filters by class/subject",
     "Student in 2+ classes",
     "1. Open history\n2. Apply filter",
     "Filtered results only show matching sessions.", "", "", "Low", ""),
    ("EH-008", "Pagination for many sessions",
     "Student has 20+ sessions",
     "1. Load history\n2. Scroll/paginate",
     "Sessions load in pages. No duplicates. Acceptable performance.", "", "", "Low", ""),
]

add_module_header(ws_tc, row, "Module 3: Exam Session History (Student)")
row += 1
style_header_row(ws_tc, row)
row += 1
for tc in module3:
    add_test_case(ws_tc, row, tc)
    row += 1
row += 1

# ── Module 4: Review Page ─────────────────────────────────────
module4 = [
    ("RV-001", "Student navigates to review page",
     "Attempt status=Graded",
     "1. From history, click 'Review'\n2. Navigate to /attempts/{attemptId}/review",
     "Review loads: exam name, attempt #, score summary, full question list.", "", "", "Critical", ""),
    ("RV-002", "Review correct/incorrect for MC question",
     "Graded attempt with MC questions",
     "1. View MC question in review",
     "Student's pick highlighted. Correct=green ✓, Wrong=red ✗.", "", "", "High", ""),
    ("RV-003", "Review correct/incorrect for TF question",
     "Graded attempt with TF questions",
     "1. View TF question in review",
     "Selection shown. Correct marked. Status badge displayed.", "", "", "High", ""),
    ("RV-004", "Review correct/incorrect for FillInBlank",
     "Graded attempt with FillInBlank questions",
     "1. View FillInBlank in review",
     "Typed answer shown vs accepted answers. explanation displayed if provided.", "", "", "High", ""),
    ("RV-005", "Review partial credit for MultipleAnswer",
     "Graded MA question; allowPartialScoring=true",
     "1. View MA question (student got some right)",
     "⚠️ Partially Correct (yellow). isPartiallyCorrect=true. Partial studentScore.", "", "", "High", ""),
    ("RV-006", "AI feedback displayed for Essay",
     "Graded attempt with Essay question",
     "1. View Essay in review",
     "essayStudentText shows response. essayAiFeedback shows AI feedback. Score shown.", "", "", "Critical", ""),
    ("RV-007", "Answer explanations shown",
     "Questions have explanation field",
     "1. View reviewed question",
     "Each answer's explanation text displayed for learning.", "", "", "Medium", ""),
    ("RV-008", "Mixed question types render correctly",
     "Attempt has all 5 question types",
     "1. Open review for mixed exam",
     "Each type renders with correct review component. No errors. Type indicated.", "", "", "High", ""),
    ("RV-009", "Score summary at top of review",
     "Any graded attempt",
     "1. Open review page",
     "Header: exam name, attempt #, total score/points, %, date, time taken.", "", "", "Medium", ""),
    ("RV-010", "Review for submitted (not yet graded) attempt",
     "Attempt status=Submitted (essay pending)",
     "1. Open review",
     "Essay shows 'Grading in progress…'. Auto-graded questions show results.", "", "", "Medium", ""),
    ("RV-011", "Access control — cannot review other student's attempt",
     "Student A tries Student B's attempt URL",
     "1. Navigate to /attempts/{otherStudentId}/review",
     "403 Forbidden. Access denied or redirect.", "", "", "Critical", ""),
    ("RV-012", "AI Explain feature on incorrect question",
     "Student viewing wrong answer in review",
     "1. Click 'Explain' on incorrect question\n2. POST /api/ai/explain",
     "Child-friendly explanation returned and displayed inline/popover.", "", "", "Medium", ""),
]

add_module_header(ws_tc, row, "Module 4: Review Page (Answer Review & Feedback)")
row += 1
style_header_row(ws_tc, row)
row += 1
for tc in module4:
    add_test_case(ws_tc, row, tc)
    row += 1
row += 1

# ── Module 5: Class Creation & Enrollment ─────────────────────
module5 = [
    ("CL-001", "Teacher creates a new class",
     "Teacher authenticated",
     "1. Navigate to classes\n2. Click 'Create Class'\n3. Fill: name, description, grade, maxStudents\n4. Submit (POST /classes)",
     "Class created: auto inviteCode (6-digit), isActive=true. Response has classId, name, grade, inviteCode.", "", "", "Critical", ""),
    ("CL-002", "Created class appears in teacher's list",
     "Class just created",
     "1. Navigate to /app/classes\n2. GET /classes?role=Teacher",
     "New class visible: name, grade, inviteCode, studentCount=0.", "", "", "High", ""),
    ("CL-003", "Class detail shows roster and assignments",
     "Class with enrolled students",
     "1. Click on class\n2. GET /classes/{id}",
     "Detail page: enrollments[] with names/avatars/status, assignments[] with exam/dates.", "", "", "High", ""),
    ("CL-004", "Student joins class with valid invite code",
     "Student authenticated; class active; not full",
     "1. Navigate to 'Join Class'\n2. Enter 6-digit code\n3. POST /classes/join",
     "Enrollment: status=Active. Response: {classId, className}. Class in student's list.", "", "", "Critical", ""),
    ("CL-005", "Student joins — invalid invite code",
     "Non-existent code",
     "1. Enter invalid code\n2. Submit",
     "Error 404/400: 'Invalid invite code'. No enrollment.", "", "", "High", ""),
    ("CL-006", "Student joins — class full",
     "studentCount >= maxStudents",
     "1. Enter valid code for full class\n2. Submit",
     "Error 400/409: 'Class is full'.", "", "", "High", ""),
    ("CL-007", "Student joins — already enrolled",
     "Student already Active in class",
     "1. Enter same invite code again",
     "Error 409: 'Already enrolled'. No duplicate.", "", "", "Medium", ""),
    ("CL-008", "Student joins — class inactive",
     "isActive=false",
     "1. Enter code for deactivated class",
     "Error 400: 'Class is not active'.", "", "", "Medium", ""),
    ("CL-009", "Teacher regenerates invite code",
     "Teacher owns the class",
     "1. Open class detail\n2. Click 'Regenerate Code'",
     "New 6-digit code generated. Old code invalid. UI updates.", "", "", "Medium", ""),
    ("CL-010", "Teacher kicks a student",
     "Teacher owns class; student enrolled",
     "1. Open roster\n2. Click 'Kick' on student\n3. Confirm",
     "Enrollment status → Kicked. Student removed from active roster.", "", "", "High", ""),
    ("CL-011", "Enrollment statuses render correctly",
     "Various enrollment states",
     "1. Query class detail",
     "Active, Inactive, Kicked, Withdrawn each with correct badge.", "", "", "Medium", ""),
    ("CL-012", "Student views enrolled classes",
     "Student in 2+ classes",
     "1. GET /classes?role=Student",
     "Only Active enrollment classes returned.", "", "", "High", ""),
    ("CL-013", "Teacher creates assignment for class",
     "Class exists; published exam exists",
     "1. Open class\n2. Add Assignment\n3. Select exam, dates, weight\n4. Submit",
     "Assignment links examId to classId. Visible in detail.", "", "", "High", ""),
    ("CL-014", "Class creation form validation",
     "Teacher on Create Class form",
     "1. Leave name empty\n2. maxStudents=0\n3. Submit",
     "Errors: 'Name required', 'Max students >= 1'. Submit blocked.", "", "", "Medium", ""),
    ("CL-015", "Admin views all classes",
     "Admin authenticated",
     "1. GET /classes?role=Admin",
     "All classes across all teachers returned.", "", "", "Medium", ""),
]

add_module_header(ws_tc, row, "Module 5: Class Creation & Student Enrollment")
row += 1
style_header_row(ws_tc, row)
row += 1
for tc in module5:
    add_test_case(ws_tc, row, tc)
    row += 1
row += 1

# ── Edge Cases & Negative Tests ───────────────────────────────
edge_cases = [
    ("EDGE-001", "Network failure during exam submission",
     "Student mid-exam; network drops",
     "1. Answer questions\n2. Disconnect network\n3. Click Submit",
     "Error toast / retry prompt. Answers cached locally. Retry on reconnect.", "", "", "Critical", ""),
    ("EDGE-002", "Concurrent session in two tabs",
     "Same session in 2 browser tabs",
     "1. Open in Tab A & B\n2. Submit from Tab A",
     "Tab B detects submission. Server rejects duplicate.", "", "", "High", ""),
    ("EDGE-003", "AI service unavailable during essay grading",
     "AI service down",
     "1. Submit exam with essay\n2. AI call fails",
     "Essay stored. Status remains Submitted. Retry/fallback available.", "", "", "High", ""),
    ("EDGE-004", "Very long essay response (10k+ chars)",
     "Student types extremely long essay",
     "1. Type 10k+ chars\n2. Submit",
     "Server accepts or returns 400 with length error. No crash.", "", "", "Medium", ""),
    ("EDGE-005", "XSS in FillInBlank answer",
     "Student types <script> tag",
     "1. Enter '<script>alert(1)</script>'\n2. Submit",
     "Input sanitized. No XSS. Stored as plain text.", "", "", "Critical", ""),
    ("EDGE-006", "Class creation boundary values",
     "Testing min/max field lengths",
     "1. name=200 chars, maxStudents=999\n2. name=1 char",
     "Long valid. Short rejected if min>1. Server validation enforced.", "", "", "Low", ""),
]

add_module_header(ws_tc, row, "Edge Cases & Negative Tests")
row += 1
style_header_row(ws_tc, row)
row += 1
for tc in edge_cases:
    add_test_case(ws_tc, row, tc)
    row += 1

# Freeze panes (freeze below first module header row won't work cleanly,
# so just freeze row 2 for scrolling with title visible)
ws_tc.freeze_panes = "A10"

# ═══════════════════════════════════════════════════════════════
#  SHEET 2 – Summary Dashboard
# ═══════════════════════════════════════════════════════════════
ws_sum = wb.create_sheet("Summary")

ws_sum.merge_cells("A1:D1")
ws_sum.cell(row=1, column=1, value="Test Execution Summary").font = TITLE_FONT

summary_headers = ["Module", "Total", "Pass", "Fail", "Blocked", "N/E", "Pass Rate (%)"]
for ci, h in enumerate(summary_headers, 1):
    c = ws_sum.cell(row=3, column=ci, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER_ALIGN
    c.border = THIN_BORDER

modules_summary = [
    ("1. Student Exam Taking Flow", len(module1)),
    ("2. Exam Generation Flow", len(module2)),
    ("3. Exam Session History", len(module3)),
    ("4. Review Page", len(module4)),
    ("5. Class Creation & Enrollment", len(module5)),
    ("Edge Cases & Negative Tests", len(edge_cases)),
]

for ri, (name, total) in enumerate(modules_summary, 4):
    ws_sum.cell(row=ri, column=1, value=name).font = NORMAL_FONT
    ws_sum.cell(row=ri, column=2, value=total).font = NORMAL_FONT
    for ci in range(1, len(summary_headers) + 1):
        ws_sum.cell(row=ri, column=ci).border = THIN_BORDER
        ws_sum.cell(row=ri, column=ci).alignment = CENTER_ALIGN

total_row = 4 + len(modules_summary)
ws_sum.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
total_cases = sum(t for _, t in modules_summary)
ws_sum.cell(row=total_row, column=2, value=total_cases).font = BOLD_FONT
for ci in range(1, len(summary_headers) + 1):
    ws_sum.cell(row=total_row, column=ci).border = THIN_BORDER
    ws_sum.cell(row=total_row, column=ci).alignment = CENTER_ALIGN

for ci in range(1, len(summary_headers) + 1):
    ws_sum.column_dimensions[get_column_letter(ci)].width = 30 if ci == 1 else 14

# ═══════════════════════════════════════════════════════════════
#  SHEET 3 – Legend
# ═══════════════════════════════════════════════════════════════
ws_leg = wb.create_sheet("Legend")

ws_leg.cell(row=1, column=1, value="Status Legend").font = TITLE_FONT
statuses = [
    ("Pass", "Test executed, result matches expected outcome"),
    ("Fail", "Test executed, result does NOT match expected outcome"),
    ("Blocked", "Cannot execute due to dependency/environment issue"),
    ("N/E", "Not Executed yet"),
]
for ri, (s, d) in enumerate(statuses, 3):
    ws_leg.cell(row=ri, column=1, value=s).font = BOLD_FONT
    ws_leg.cell(row=ri, column=2, value=d).font = NORMAL_FONT

ws_leg.cell(row=8, column=1, value="Severity Legend").font = TITLE_FONT
severities = [
    ("Critical", "Blocks core functionality, no workaround", "FF4444"),
    ("High", "Major feature impacted, workaround possible", "FFA500"),
    ("Medium", "Minor feature impacted, easy workaround", "FFD700"),
    ("Low", "Cosmetic or minor UX issue", "90EE90"),
]
for ri, (s, d, color) in enumerate(severities, 10):
    c = ws_leg.cell(row=ri, column=1, value=s)
    c.font = BOLD_FONT
    c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws_leg.cell(row=ri, column=2, value=d).font = NORMAL_FONT

ws_leg.column_dimensions["A"].width = 15
ws_leg.column_dimensions["B"].width = 55

# ═══════════════════════════════════════════════════════════════
#  Save
# ═══════════════════════════════════════════════════════════════
output_path = "FrogEdu_Test_Report.xlsx"
wb.save(output_path)
print(f"✅ Generated {output_path} with {total_cases} test cases across 3 sheets.")
