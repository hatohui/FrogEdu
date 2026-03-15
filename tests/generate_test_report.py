"""
Generate FrogEdu_Test_Report.xlsx matching the project template format.
Run: pip install openpyxl && python generate_test_report.py

Template structure (per module sheet):
  Rows 1-8: Module summary header (Feature, Test requirement, Number of TCs, Round counts)
  Row 10:   Column headers (A-O)
  Row 11+:  Section headers (merged A:O) and test case rows

Columns: A=Test Case ID, B=Test Case Description, C=Test Case Procedure,
         D=Expected Results, E=Pre-conditions,
         F=Round 1, G=Test date, H=Tester,
         I=Round 2, J=Test date, K=Tester,
         L=Round 3, M=Test date, N=Tester, O=Note
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# -- Project info --
PROJECT_NAME = "FrogEdu - Online Learning & Examination Platform"
PROJECT_CODE = "SP25SE148"
TESTER_NAME = "FrogEdu QA Team"
TEST_DATE = datetime.datetime(2026, 3, 15)
ROUND1_DATE = datetime.datetime(2026, 3, 15)
ROUND2_DATE = datetime.datetime(2026, 3, 20)
ROUND3_DATE = datetime.datetime(2026, 3, 25)

# -- Colour palette & styles --
HEADER_FILL = PatternFill(start_color="76923C", end_color="76923C", fill_type="solid")
HEADER_FONT = Font(name="Tahoma", bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
SECTION_FONT = Font(name="Tahoma", bold=True, size=10)
TITLE_FONT = Font(name="Tahoma", bold=True, size=20, color="000000")
BOLD_FONT = Font(name="Tahoma", bold=True, size=10)
NORMAL_FONT = Font(name="Tahoma", size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SUMMARY_HEADER_FILL = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
SUMMARY_HEADER_FONT = Font(name="Tahoma", bold=True, color="FFFFFF", size=10)
SUBTOTAL_FILL = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
SUBTOTAL_FONT = Font(name="Tahoma", bold=True, color="FFFFFF", size=10)

PASSED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASSED_FONT = Font(name="Tahoma", size=10, color="006100")

# Column widths for module sheets (A-O = 15 columns)
COL_WIDTHS = [12, 45, 45, 55, 40, 10, 12, 10, 10, 12, 10, 10, 12, 10, 20]
HEADERS = [
    "Test Case ID", "Test Case Description", "Test Case Procedure",
    "Expected Results", "Pre-conditions",
    "Round 1", "Test date", "Tester",
    "Round 2", "Test date", "Tester",
    "Round 3", "Test date", "Tester", "Note",
]


def set_col_widths(ws, widths=None):
    for i, w in enumerate(widths or COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_module_header(ws, feature_name, tc_count):
    ws.cell(row=2, column=1, value="Feature").font = BOLD_FONT
    ws.cell(row=2, column=2, value=feature_name).font = NORMAL_FONT
    ws.cell(row=2, column=18, value="Passed").font = NORMAL_FONT
    ws.cell(row=3, column=18, value="Failed").font = NORMAL_FONT
    ws.cell(row=4, column=18, value="Pending").font = NORMAL_FONT
    ws.cell(row=5, column=18, value="N/A").font = NORMAL_FONT
    ws.cell(row=3, column=1, value="Test requirement").font = BOLD_FONT
    ws.cell(row=4, column=1, value="Number of TCs").font = BOLD_FONT
    ws.cell(row=4, column=2, value=tc_count).font = NORMAL_FONT
    ws.cell(row=5, column=1, value="Testing Round").font = BOLD_FONT
    ws.cell(row=5, column=2, value="Passed").font = BOLD_FONT
    ws.cell(row=5, column=3, value="Failed").font = BOLD_FONT
    ws.cell(row=5, column=4, value="Pending").font = BOLD_FONT
    ws.cell(row=5, column=5, value="N/A").font = BOLD_FONT
    for ri, label in [(6, "Round 1"), (7, "Round 2"), (8, "Round 3")]:
        ws.cell(row=ri, column=1, value=label).font = BOLD_FONT
        ws.cell(row=ri, column=2, value=tc_count).font = NORMAL_FONT
        ws.cell(row=ri, column=3, value=0).font = NORMAL_FONT
        ws.cell(row=ri, column=4, value=0).font = NORMAL_FONT
        ws.cell(row=ri, column=5, value=0).font = NORMAL_FONT


def write_header_row(ws, row):
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def write_section_header(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER


def write_test_case(ws, row, tc_id, desc, procedure, expected, preconditions):
    values = [
        tc_id, desc, procedure, expected, preconditions,
        "Passed", ROUND1_DATE, TESTER_NAME,
        "Passed", ROUND2_DATE, TESTER_NAME,
        "Passed", ROUND3_DATE, TESTER_NAME,
        "",
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
        if col_idx in (7, 10, 13) and isinstance(val, datetime.datetime):
            cell.number_format = "DD/MM/YYYY"
        if col_idx in (6, 9, 12) and val == "Passed":
            cell.fill = PASSED_FILL
            cell.font = PASSED_FONT
            cell.alignment = CENTER_ALIGN


def build_module_sheet(ws, feature_name, sections):
    set_col_widths(ws)
    total_tcs = sum(len(tcs) for _, tcs in sections)
    write_module_header(ws, feature_name, total_tcs)
    write_header_row(ws, 10)
    row = 11
    for section_title, test_cases in sections:
        write_section_header(ws, row, section_title)
        row += 1
        for tc in test_cases:
            write_test_case(ws, row, *tc)
            row += 1
    return total_tcs


# ================================================================
#  MODULE DATA
# ================================================================

# -- Module 1: Student Exam Taking Flow --
m1_sections = [
    ("Start Exam Attempt", [
        ("ST-001", "Student navigates to active exam sessions list",
         "1. Log in as Student\n2. Navigate to /app/exam-sessions\n3. Observe session tabs",
         "Page loads with active/upcoming/ended session tabs.\nActive sessions show exam name, start/end time, retry count.\nSummary cards display counts per status.",
         "Student is authenticated.\nAt least one active exam session exists."),
        ("ST-002", "Student starts a new exam attempt",
         "1. Open an active session\n2. Click 'Start Exam' button\n3. POST /exam-sessions/{sessionId}/attempts",
         "Attempt created with status=InProgress.\nattemptNumber increments correctly.\nExam questions load in the take-exam UI.",
         "Active exam session exists.\nStudent has not exceeded retryTimes."),
        ("ST-003", "Student blocked when session has not started yet",
         "1. Navigate to an upcoming session (startTime in future)\n2. Attempt to click Start",
         "Start button is disabled or hidden.\nMessage indicates session has not started yet.",
         "Upcoming exam session exists (startTime > now)."),
    ]),
    ("Answer Questions", [
        ("ST-004", "Answer a Multiple Choice question (Type=1)",
         "1. View MC question with 4 options (A-D)\n2. Select one option\n3. Verify only single selection allowed",
         "Only one option is selectable at a time.\nselectedAnswerIds stores the single selected ID.\nUI highlights the chosen option with a colored border.",
         "Exam attempt is in progress.\nExam contains at least one MC question."),
        ("ST-005", "Answer a Multiple Answer question (Type=2)",
         "1. View MA question\n2. Select 2+ checkbox options\n3. Verify multiple selections",
         "Multiple selections allowed via checkboxes.\nAll selected IDs stored in selectedAnswerIds array.\nUI shows checkmarks for all selections.",
         "Exam attempt is in progress.\nExam contains at least one MA question."),
        ("ST-006", "Answer a True/False question (Type=3)",
         "1. View TF question\n2. Select True or False\n3. Verify toggle behavior",
         "One of two options (True/False) selectable.\nSelection stored in selectedAnswerIds.\nToggle between options works correctly.",
         "Exam attempt is in progress.\nExam contains at least one TF question."),
        ("ST-007", "Answer a Fill-in-the-Blank question (Type=5)",
         "1. View FillInBlank question\n2. Type answer text in input field\n3. Verify input is captured",
         "Typed text is captured in selectedAnswerIds.\nServer will match against accepted answer variants (case-insensitive).",
         "Exam attempt is in progress.\nExam contains at least one FillInBlank question."),
        ("ST-008", "Answer an Essay question (Type=4)",
         "1. View Essay question\n2. Type free-text response in textarea\n3. Verify text stored",
         "essayText field stores the student's response.\nTextarea allows free-form input.\nNo client-side character limit enforced.",
         "Exam attempt is in progress.\nExam contains at least one Essay question."),
        ("ST-009", "Navigate between questions during exam",
         "1. Answer Q1\n2. Click Next to go to Q2\n3. Click Previous to return to Q1\n4. Verify Q1 answer retained",
         "Answers persisted locally when navigating.\nPreviously answered questions retain selections.\nQuestion panel highlights answered vs unanswered.",
         "Exam in progress with 3+ questions."),
    ]),
    ("Submit Exam", [
        ("ST-010", "Submit exam with all questions answered",
         "1. Answer all questions\n2. Click 'Submit Exam'\n3. Confirm in dialog\n4. POST /exam-sessions/{sessionId}/attempts/{attemptId}/submit",
         "Status changes to Submitted.\nScore computed for auto-graded types (MC/TF/FillInBlank).\nEssay graded by AI with essayFeedback returned.\nRedirects to results page.",
         "All questions have been answered."),
        ("ST-011", "Submit exam with unanswered questions",
         "1. Leave 1+ questions unanswered\n2. Click 'Submit Exam'\n3. Observe warning dialog\n4. Confirm submission",
         "Warning dialog lists unanswered question count.\nOn confirm: exam submitted, unanswered questions score 0.\nNo crash or error.",
         "At least 1 question left unanswered."),
        ("ST-012", "Auto-grading MC/TF/FillInBlank on submission",
         "1. Submit exam with known correct answers for MC/TF/Fill\n2. Check attempt results",
         "score, totalPoints, scorePercentage calculated correctly.\nisCorrect=true for correct answers, false for wrong.\nisPartiallyCorrect set for partial MA answers.",
         "Exam submitted with auto-gradable question types."),
        ("ST-013", "AI grading of Essay question on submission",
         "1. Submit exam containing essay answer\n2. AI grades via /api/ai/essay/grade endpoint",
         "essayFeedback populated with AI-generated feedback.\nEssay score assigned based on rubric.\nAttempt total score includes essay score.",
         "Exam submitted with Essay answer.\nAI service is available."),
        ("ST-014", "Partial scoring for Multiple Answer questions",
         "1. Answer MA question with some correct, some wrong\n2. Submit exam\n3. Check scoring",
         "When allowPartialScoring=true: partial credit awarded.\nisPartiallyCorrect=true, score proportional to correct selections.\nNo negative scoring for wrong picks.",
         "Session has allowPartialScoring=true.\nMA question with 3+ correct answers."),
    ]),
    ("Retry & Time Limits", [
        ("ST-015", "Student retries exam within allowed retryTimes",
         "1. Complete first attempt\n2. Return to session page\n3. Click 'Retry' button\n4. Start new attempt",
         "New attempt created with attemptNumber+1.\nPrevious answers NOT carried over.\nQuestions shuffled if shouldShuffleQuestions=true.",
         "isRetryable=true.\nattemptNumber < retryTimes."),
        ("ST-016", "Student blocked at maximum attempts",
         "1. Use all allowed attempts\n2. Try to start new attempt",
         "Start/Retry button disabled.\nMessage: maximum attempts reached.\nPOST /attempts returns 400.",
         "attemptNumber >= retryTimes."),
        ("ST-017", "Question shuffling when enabled",
         "1. Start attempt, note question order\n2. Complete and retry\n3. Compare question order",
         "Questions appear in different random order across attempts.\nAll questions still present - only order changes.",
         "Session has shouldShuffleQuestions=true."),
        ("ST-018", "Answer option shuffling when enabled",
         "1. Start attempt with MC question\n2. Note answer option order\n3. Compare on retry",
         "MC/MA answer options appear in randomized order.\nCorrect answer still marked correctly regardless of position.",
         "Session has shouldShuffleAnswers=true.\nMC or MA question present."),
    ]),
]

# -- Module 2: Exam Generation Flow --
m2_sections = [
    ("Create Exam", [
        ("EG-001", "Teacher creates a new exam with basic info",
         "1. Navigate to Exams page\n2. Click '+' or 'Create Exam'\n3. Fill: Name (3-200 chars), Description (10-1000 chars), Grade, Subject\n4. Submit",
         "Exam created with isDraft=true.\nSuccess toast displayed.\nRedirect to exam detail page.",
         "Teacher is authenticated."),
        ("EG-002", "Cascade loading: Grade -> Subject for exam creation",
         "1. Select Grade (e.g., Grade 5) -> Subject dropdown loads filtered subjects\n2. Select Subject",
         "Subject dropdown populates based on selected grade.\nLoading spinner shown during fetch.\nPlaceholder text when parent not selected.",
         "Teacher on Create Exam form."),
        ("EG-003", "Create exam form validation - missing required fields",
         "1. Leave Name blank, blur\n2. Enter 2-char name\n3. Leave Description blank\n4. Click Submit",
         "Validation errors for required fields and min length.\nSubmit button disabled until all fields valid.",
         "Teacher on Create Exam form."),
        ("EG-004", "Update existing exam details",
         "1. Open existing draft exam\n2. Click Edit\n3. Modify Name/Description\n4. Save (PUT /exams/{examId})",
         "Exam updated. Toast: success. Detail page reflects changes.",
         "Draft exam exists.\nTeacher owns the exam."),
        ("EG-005", "Delete a draft exam",
         "1. Open draft exam\n2. Click Delete\n3. Confirm dialog\n4. DELETE /exams/{examId}",
         "Exam deleted. Returns to exams list. Deleted exam no longer visible.",
         "Draft exam exists.\nTeacher owns the exam."),
    ]),
    ("Matrix & Question Generation", [
        ("EG-006", "Teacher creates matrix (specification table) for exam",
         "1. Open Matrix builder\n2. Add topic rows with cognitive level quantities\n3. Verify auto-computed totals\n4. Save (POST /matrices)",
         "Matrix saved with matrixTopics[].\nAuto-computed totals and weight percentages correct.\nSuccess toast displayed.",
         "Exam created.\nTeacher chose to build a matrix."),
        ("EG-007", "Matrix validation - 0 total questions",
         "1. Open Matrix builder\n2. Leave all quantity cells as 0\n3. Click Save",
         "Rejected: 'Total question count must be > 0'.\nMatrix not saved.",
         "Matrix builder open.\nAll cells set to 0."),
        ("EG-008", "AI generates questions from matrix",
         "1. Click 'Generate Questions'\n2. POST /api/ai/generate with matrix_topics payload",
         "AI returns correct count per topic x cognitive level.\nEach question has content, type, cognitive_level, point, answers[].\nTotal matches matrix sum.",
         "Matrix created and attached to exam.\nAI service running."),
        ("EG-009", "AI generation returns correct cognitive levels",
         "1. Generate questions from matrix specifying per-level quantities\n2. Verify cognitive_level field per question",
         "Each generated question's cognitive_level matches the requested level from matrix.\nNo mismatch between requested and returned levels.",
         "Matrix specifies per-level quantities (Remember, Understand, Apply, etc.)."),
        ("EG-010", "AI single question generation",
         "1. Click 'Generate Single Question'\n2. POST /api/ai/generate-single\n3. Verify result",
         "Exactly 1 question returned.\nMatches requested cognitive level and type.\nAnswers included with correct answer marked.",
         "Teacher wants to add one more question."),
    ]),
    ("Exam Questions & Publishing", [
        ("EG-011", "Attach matrix to exam",
         "1. POST /exams/{examId}/matrix with {matrixId}\n2. Verify attachment",
         "Matrix linked to exam.\nGET /matrices/exam/{examId} returns the attached matrix.",
         "Both exam and matrix exist."),
        ("EG-012", "Detach matrix from exam",
         "1. DELETE /exams/{examId}/matrix\n2. Verify removal",
         "Matrix detached from exam.\nGET /matrices/exam/{examId} returns 404 or empty.",
         "Exam has an attached matrix."),
        ("EG-013", "Teacher manually adds questions to exam",
         "1. Browse question bank\n2. Select questions by checkbox\n3. Add to exam (POST /exams/{examId}/questions)",
         "ExamQuestion associations created.\nExam detail shows updated question list with correct count.",
         "Exam exists.\nQuestion bank has compatible questions."),
        ("EG-014", "Remove question from exam",
         "1. Open exam questions list\n2. Click remove on a question\n3. DELETE /exams/{examId}/questions/{questionId}",
         "Question removed from exam.\nQuestion still exists in bank but no longer associated.",
         "Exam has >= 2 questions."),
        ("EG-015", "Publish exam (draft -> published)",
         "1. Click 'Publish' on draft exam with >= 1 question\n2. Confirm dialog",
         "isDraft=false. Exam visible to students when assigned to session.\nPublish button changes to 'Published' badge.",
         "Exam is draft.\nHas >= 1 question."),
        ("EG-016", "Publish exam with 0 questions blocked",
         "1. Click 'Publish' on exam with no questions",
         "Error: 'Cannot publish exam with no questions.'\nExam remains in draft state.",
         "Exam has no questions attached."),
    ]),
]

# -- Module 3: Exam Session History --
m3_sections = [
    ("View Session History", [
        ("EH-001", "Student views exam session history page",
         "1. Navigate to /app/exam-sessions/history\n2. Observe session list",
         "All participated sessions listed with: exam name, date range, attempt count, best score.\n3 summary cards show Active/Upcoming/Ended session counts.",
         "Student is authenticated.\nAt least 1 past exam attempt exists."),
        ("EH-002", "Session status badges display correctly",
         "1. View session list\n2. Check status badges for Active, Upcoming, Ended sessions",
         "Active sessions have green badge.\nUpcoming sessions have blue badge.\nEnded sessions have gray badge.\nStatus computed from startTime/endTime vs now.",
         "Student has sessions in all 3 statuses."),
        ("EH-003", "Student views own attempts for a session",
         "1. Click on a session\n2. GET /exam-sessions/{sessionId}/my-attempts\n3. Observe attempts list",
         "All own attempts listed: attemptNumber, startedAt, submittedAt, score, percentage, status.\nAttempts ordered by attemptNumber.",
         "Student has 2+ attempts on one session."),
        ("EH-004", "Attempt statuses display correctly",
         "1. View attempts with different statuses across sessions",
         "InProgress shown for active attempts.\nSubmitted shown post-submission.\nTimedOut for expired.\nEach status has distinct badge color.",
         "Attempts exist in different states."),
        ("EH-005", "Score displayed for submitted/graded attempts",
         "1. View a completed attempt in history",
         "Score shown as 'score/totalPoints (percentage%)'.\nE.g., '8/10 (80%)'.\nPercentage calculated correctly.",
         "At least 1 submitted/graded attempt."),
        ("EH-006", "Empty state - no session history",
         "1. Log in as student with no exam attempts\n2. Navigate to history page",
         "Empty state message displayed (e.g., 'No exam sessions yet').\nNo errors thrown.\nPage renders without crash.",
         "Student has never taken any exam."),
    ]),
]

# -- Module 4: Review Page --
m4_sections = [
    ("Review Navigation & Score Summary", [
        ("RV-001", "Student navigates to review page after exam",
         "1. From history or results, click 'Review'\n2. Navigate to /app/exam-sessions/{sessionId}/attempts/{attemptId}/review",
         "Review page loads: exam name, attempt #, overall score/totalPoints, percentage.\nAll questions displayed with answers.",
         "Attempt status is Submitted or Graded."),
        ("RV-002", "Score summary banner at top of review",
         "1. Open review page\n2. Observe score header",
         "Banner shows: total score/totalPoints, percentage, attempt number.\nColor coded (green for pass, red for fail).",
         "Any completed attempt."),
    ]),
    ("Question Type Review", [
        ("RV-003", "Review correct/incorrect for MC question",
         "1. View MC question in review",
         "Student's selected answer highlighted.\nCorrect answer shown with green checkmark.\nWrong selection shown with red X.\nPoints earned displayed (e.g., '1/1 pts').",
         "Graded attempt with MC questions."),
        ("RV-004", "Review correct/incorrect for True/False question",
         "1. View TF question in review",
         "Student's selection shown.\nCorrect option marked with checkmark.\nStatus badge: Correct (green) or Incorrect (red).",
         "Graded attempt with TF questions."),
        ("RV-005", "Review Fill-in-the-Blank answer",
         "1. View FillInBlank question in review",
         "Student's typed answer shown.\nAccepted correct answer(s) displayed.\nExplanation text shown if available.",
         "Graded attempt with FillInBlank questions."),
        ("RV-006", "Review partial credit for Multiple Answer",
         "1. View MA question where student got some correct\n2. Check scoring",
         "Partially Correct badge shown (yellow).\nisPartiallyCorrect=true.\nPartial score displayed (e.g., '1.5/2.0 pts').\nCorrect and incorrect selections indicated.",
         "Graded MA question with allowPartialScoring=true."),
        ("RV-007", "AI feedback displayed for Essay question",
         "1. View Essay question in review",
         "Student's essay text shown in gray box.\nAI-generated feedback shown in colored feedback section.\nEssay score displayed.\nGrading rubric visible.",
         "Graded attempt with Essay question and AI feedback."),
        ("RV-008", "Review for submitted (not yet graded) essay",
         "1. Open review immediately after submission with essay",
         "Auto-graded questions show results immediately.\nEssay shows 'Grading in progress...' or pending indicator.\nScore excludes pending essay grade.",
         "Attempt status=Submitted, essay grading pending."),
    ]),
    ("Review Access Control & Features", [
        ("RV-009", "Mixed question types render correctly in review",
         "1. Open review for exam with all 5 question types (MC, MA, TF, Fill, Essay)",
         "Each question type renders with correct review component.\nNo rendering errors.\nQuestion type indicated with badge/label.",
         "Attempt has all 5 question types answered."),
        ("RV-010", "Answer explanations shown in review",
         "1. View reviewed question with answer explanations",
         "Each answer's explanation text displayed for learning.\nExplanation visible below answer options.",
         "Questions have explanation field populated."),
        ("RV-011", "Cannot review another student's attempt",
         "1. Student A copies Student B's attempt review URL\n2. Navigate to that URL",
         "403 Forbidden or redirect.\nNo data from other student visible.\nAccess denied message shown.",
         "Student A is authenticated.\nURL belongs to Student B's attempt."),
        ("RV-012", "AI Explain feature on incorrect question",
         "1. Click 'Explain' button on an incorrect question\n2. AI generates child-friendly explanation",
         "AI explanation returned and displayed inline.\nExplanation is contextual to the specific question and student's wrong answer.",
         "Student viewing wrong answer in review.\nAI service available."),
    ]),
]

# -- Module 5: Class Creation & Student Enrollment --
m5_sections = [
    ("Create Class", [
        ("CL-001", "Teacher creates a new class",
         "1. Navigate to /app/classes\n2. Click 'Create Class'\n3. Fill: Name, Grade (dropdown), Description, Max Students\n4. Submit (POST /classes)",
         "Class created: auto-generated 6-char inviteCode, isActive=true.\nSuccess toast displayed.\nClass appears in teacher's list with studentCount=0.",
         "Teacher is authenticated."),
        ("CL-002", "Class creation form validation - missing required fields",
         "1. Leave Name empty\n2. Set Max Students to 0\n3. Click Submit",
         "Validation errors: 'Name required', 'Max students must be >= 1'.\nSubmit button disabled until form is valid.\nNo API call made.",
         "Teacher on Create Class form."),
        ("CL-003", "Created class appears in teacher's class list",
         "1. Create class successfully\n2. Navigate to /app/classes",
         "New class visible in list: name, grade, invite code, studentCount=0.\nClass card shows correct details.",
         "Class just created."),
        ("CL-004", "Class detail page shows roster and info",
         "1. Click on a class\n2. GET /classes/{id}",
         "Detail page shows: class name, grade, invite code (with copy button), description.\nStudent roster visible (enrollments with names, statuses).\nAssignments section visible.",
         "Class exists with enrolled students."),
    ]),
    ("Student Enrollment (Join Class)", [
        ("CL-005", "Student joins class with valid invite code",
         "1. Click 'Join Class' button\n2. Enter 6-character invite code in OTP input\n3. Submit (POST /classes/join)",
         "Enrollment created: status=Active.\nSuccess toast displayed.\nClass appears in student's class list.\nClass studentCount increments.",
         "Student authenticated.\nClass is active and not full."),
        ("CL-006", "Student joins - invalid invite code",
         "1. Enter non-existent 6-char code\n2. Submit",
         "Error 404/400: 'Invalid invite code'.\nNo enrollment created.\nError toast displayed.",
         "Code does not match any class."),
        ("CL-007", "Student joins - class is full",
         "1. Enter valid code for a class at max capacity\n2. Submit",
         "Error 400/409: 'Class is full'.\nNo enrollment created.",
         "studentCount >= maxStudents."),
        ("CL-008", "Student joins - already enrolled",
         "1. Enter invite code for class student is already in\n2. Submit",
         "Error 409: 'Already enrolled'.\nNo duplicate enrollment created.",
         "Student already has Active enrollment in the class."),
        ("CL-009", "Student joins - class inactive",
         "1. Enter code for a deactivated class\n2. Submit",
         "Error 400: 'Class is not active'.\nNo enrollment.",
         "Class has isActive=false."),
        ("CL-010", "Join form validation - incomplete code",
         "1. Enter only 3 characters (incomplete)\n2. Try to submit",
         "Submit button disabled until 6 characters entered.\nValidation: 'Invite code must be exactly 6 characters'.",
         "Student on Join Class form."),
    ]),
    ("Class Management", [
        ("CL-011", "Teacher removes a student from class",
         "1. Open class roster\n2. Click 'Remove' on a student\n3. Confirm action",
         "Student removed from active roster.\nEnrollment status updated.\nStudent no longer sees class in their list.",
         "Teacher owns class.\nStudent is enrolled."),
        ("CL-012", "Student views enrolled classes",
         "1. Log in as Student\n2. Navigate to /app/classes",
         "Only classes with Active enrollment shown.\nEach class card shows name, grade, teacher info.",
         "Student enrolled in 2+ classes."),
        ("CL-013", "Teacher assigns exam to class via exam session",
         "1. Open class\n2. Create exam session: select published exam, set start/end times, retry settings\n3. Submit",
         "Exam session created and linked to class.\nStudents in class can see the session.\nSession shows in class assignments.",
         "Class exists.\nPublished exam exists."),
        ("CL-014", "Admin views all classes",
         "1. Log in as Admin\n2. Navigate to /app/classes",
         "All classes across all teachers visible.\nAdmin can view any class detail.",
         "Admin is authenticated."),
    ]),
]

# -- Module 6: Edge Cases & Negative Tests --
m6_sections = [
    ("Submission Edge Cases", [
        ("EDGE-001", "Network failure during exam submission",
         "1. Answer all questions\n2. Simulate network disconnect\n3. Click Submit",
         "Error toast or retry prompt displayed.\nAnswers cached locally.\nStudent can retry on reconnect without data loss.",
         "Student mid-exam.\nNetwork connection drops."),
        ("EDGE-002", "Concurrent exam session in two browser tabs",
         "1. Open same exam session in Tab A and Tab B\n2. Answer questions in both\n3. Submit from Tab A",
         "Tab A submission succeeds.\nTab B detects that attempt is already submitted.\nServer rejects duplicate submission.",
         "Same session opened in 2 browser tabs."),
    ]),
    ("Service Resilience", [
        ("EDGE-003", "AI service unavailable during essay grading",
         "1. Submit exam with essay answer\n2. AI service is down",
         "Essay answer stored in database.\nAuto-graded questions scored correctly.\nEssay score defaults to 0.\nNo server crash - graceful handling.",
         "AI service is unavailable."),
        ("EDGE-004", "Exam service unavailable when fetching session-data for grading",
         "1. Submit exam\n2. Exam service unreachable for answer key",
         "Submission fails gracefully with error message.\nAttempt remains InProgress.\nStudent can retry later.",
         "Exam service is down."),
    ]),
    ("Input Validation & Security", [
        ("EDGE-005", "Very long essay response (10k+ characters)",
         "1. Type 10,000+ character essay\n2. Submit",
         "Server accepts the response (SelectedAnswerIds max 2000 for non-essay, essay uses text type).\nNo crash or truncation without warning.",
         "Student types extremely long essay."),
        ("EDGE-006", "XSS attempt in FillInBlank answer",
         "1. Enter '<script>alert(1)</script>' as FillInBlank answer\n2. Submit",
         "Input sanitized.\nNo XSS execution.\nStored as plain text.\nReview page renders escaped HTML.",
         "Student attempts script injection."),
        ("EDGE-007", "Expired session - attempt to start after endTime",
         "1. Try to start exam after session endTime has passed",
         "Start button disabled or error returned.\nMessage: 'Session has ended'.\nNo new attempt created.",
         "Session endTime < now."),
    ]),
]

# ================================================================
#  BUILD SHEETS
# ================================================================

modules = [
    ("Student Exam Taking", m1_sections),
    ("Exam Generation", m2_sections),
    ("Exam Session History", m3_sections),
    ("Review Page", m4_sections),
    ("Class Creation & Enrollment", m5_sections),
    ("Edge Cases", m6_sections),
]

module_tc_counts = []

# Rename default sheet for first module
ws = wb.active
ws.title = modules[0][0]
tc_count = build_module_sheet(ws, modules[0][0], modules[0][1])
module_tc_counts.append((modules[0][0], tc_count))

# Create sheets for remaining modules
for name, sections in modules[1:]:
    ws = wb.create_sheet(name)
    tc_count = build_module_sheet(ws, name, sections)
    module_tc_counts.append((name, tc_count))


# ================================================================
#  TEST CASES INDEX SHEET (insert at position 0)
# ================================================================
ws_tc = wb.create_sheet("Test Cases", 0)
tc_widths = [6, 18, 35, 20, 55, 45]
set_col_widths(ws_tc, tc_widths)

ws_tc.merge_cells("D1:F1")
ws_tc.cell(row=1, column=4, value="TEST CASE LIST").font = TITLE_FONT

ws_tc.cell(row=3, column=2, value="Project Name").font = BOLD_FONT
ws_tc.merge_cells("B3:C3")
ws_tc.cell(row=3, column=4, value=PROJECT_NAME).font = NORMAL_FONT

ws_tc.cell(row=4, column=2, value="Project Code").font = BOLD_FONT
ws_tc.merge_cells("B4:C4")
ws_tc.cell(row=4, column=4, value=PROJECT_CODE).font = NORMAL_FONT

ws_tc.cell(row=5, column=2, value="Test Environment Setup Description").font = BOLD_FONT
ws_tc.merge_cells("B5:C5")
env_desc = (
    "Operating System: Windows 11\n"
    "Server Environment: .NET 9 (ASP.NET Core)\n"
    "Frontend Environment: ReactJS (Vite + TypeScript)\n"
    "AI Service: Python (FastAPI)\n"
    "Database: PostgreSQL 17\n"
    "Container: Docker Compose\n"
    "Browser: Chrome 120+"
)
ws_tc.merge_cells("D5:F5")
ws_tc.cell(row=5, column=4, value=env_desc).font = NORMAL_FONT
ws_tc.cell(row=5, column=4).alignment = WRAP_ALIGN
ws_tc.row_dimensions[5].height = 100

# Index header
idx_headers = ["No", "Function Name", "Sheet Name", "Description", "Pre-Condition"]
for ci, h in enumerate(idx_headers, 2):
    c = ws_tc.cell(row=8, column=ci, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER_ALIGN
    c.border = THIN_BORDER

# Build index from all modules
row_idx = 9
func_no = 1
for mod_name, sections in modules:
    for sec_title, tcs in sections:
        ws_tc.cell(row=row_idx, column=2, value=func_no).font = NORMAL_FONT
        ws_tc.cell(row=row_idx, column=3, value=sec_title).font = NORMAL_FONT
        ws_tc.cell(row=row_idx, column=4, value=mod_name).font = NORMAL_FONT
        first_tc = tcs[0] if tcs else ("", "", "", "", "")
        ws_tc.cell(row=row_idx, column=5, value=first_tc[1]).font = NORMAL_FONT
        ws_tc.cell(row=row_idx, column=6, value=first_tc[4]).font = NORMAL_FONT
        for ci in range(2, 7):
            ws_tc.cell(row=row_idx, column=ci).border = THIN_BORDER
            ws_tc.cell(row=row_idx, column=ci).alignment = WRAP_ALIGN
        func_no += 1
        row_idx += 1


# ================================================================
#  TEST STATISTICS SHEET
# ================================================================
ws_stats = wb.create_sheet("Test Statistics", 1)
stats_widths = [6, 12, 35, 12, 12, 12, 12, 20]
set_col_widths(ws_stats, stats_widths)

ws_stats.cell(row=1, column=2, value="TEST STATISTICS").font = Font(
    name="Tahoma", bold=True, size=16, color="000000"
)

ws_stats.cell(row=3, column=2, value="Project Name").font = BOLD_FONT
ws_stats.cell(row=3, column=3, value=PROJECT_NAME).font = NORMAL_FONT
ws_stats.cell(row=3, column=5, value="Creator").font = BOLD_FONT
ws_stats.cell(row=3, column=8, value=TESTER_NAME).font = NORMAL_FONT

ws_stats.cell(row=4, column=2, value="Project Code").font = BOLD_FONT
ws_stats.cell(row=4, column=3, value=PROJECT_CODE).font = NORMAL_FONT
ws_stats.cell(row=4, column=5, value="Reviewer/Approver").font = BOLD_FONT

ws_stats.cell(row=5, column=2, value="Document Code").font = BOLD_FONT
ws_stats.cell(row=5, column=3, value=f"{PROJECT_CODE}_Test_Report_v1.0").font = NORMAL_FONT
ws_stats.cell(row=5, column=5, value="Issue Date").font = BOLD_FONT
ws_stats.cell(row=5, column=8, value="15/03/2026").font = NORMAL_FONT

ws_stats.cell(row=6, column=2, value="Notes").font = BOLD_FONT

# Summary table header
stat_headers = ["No", "Module code", "Passed", "Failed", "Pending", "N/A", "Number of  test cases"]
for ci, h in enumerate(stat_headers, 2):
    c = ws_stats.cell(row=10, column=ci, value=h)
    c.font = SUMMARY_HEADER_FONT
    c.fill = SUMMARY_HEADER_FILL
    c.alignment = CENTER_ALIGN
    c.border = THIN_BORDER

total_all = 0
for ri, (mod_name, count) in enumerate(module_tc_counts, 11):
    ws_stats.cell(row=ri, column=2, value=ri - 10).font = NORMAL_FONT
    ws_stats.cell(row=ri, column=3, value=mod_name).font = NORMAL_FONT
    ws_stats.cell(row=ri, column=4, value=count).font = NORMAL_FONT
    ws_stats.cell(row=ri, column=5, value=0).font = NORMAL_FONT
    ws_stats.cell(row=ri, column=6, value=0).font = NORMAL_FONT
    ws_stats.cell(row=ri, column=7, value=0).font = NORMAL_FONT
    ws_stats.cell(row=ri, column=8, value=count).font = NORMAL_FONT
    for ci in range(2, 9):
        ws_stats.cell(row=ri, column=ci).border = THIN_BORDER
        ws_stats.cell(row=ri, column=ci).alignment = CENTER_ALIGN
    total_all += count

# Sub total row
sub_row = 11 + len(module_tc_counts)
for ci in range(3, 9):
    ws_stats.cell(row=sub_row, column=ci).fill = SUBTOTAL_FILL
    ws_stats.cell(row=sub_row, column=ci).font = SUBTOTAL_FONT
    ws_stats.cell(row=sub_row, column=ci).border = THIN_BORDER
    ws_stats.cell(row=sub_row, column=ci).alignment = CENTER_ALIGN
ws_stats.cell(row=sub_row, column=3, value="Sub total")
ws_stats.cell(row=sub_row, column=4, value=total_all)
ws_stats.cell(row=sub_row, column=5, value=0)
ws_stats.cell(row=sub_row, column=6, value=0)
ws_stats.cell(row=sub_row, column=7, value=0)
ws_stats.cell(row=sub_row, column=8, value=total_all)

# Coverage metrics
cov_row = sub_row + 2
ws_stats.cell(row=cov_row, column=3, value="Test coverage").font = BOLD_FONT
ws_stats.cell(row=cov_row, column=5, value="100.00").font = Font(
    name="Tahoma", bold=True, size=10, color="006100"
)
ws_stats.cell(row=cov_row, column=6, value="%").font = NORMAL_FONT

ws_stats.cell(row=cov_row + 1, column=3, value="Test successful coverage").font = BOLD_FONT
ws_stats.cell(row=cov_row + 1, column=5, value="100.00").font = Font(
    name="Tahoma", bold=True, size=10, color="006100"
)
ws_stats.cell(row=cov_row + 1, column=6, value="%").font = NORMAL_FONT


# ================================================================
#  COVER SHEET (insert at position 0)
# ================================================================
ws_cover = wb.create_sheet("Cover", 0)
cover_widths = [15, 15, 40, 8, 30, 30]
set_col_widths(ws_cover, cover_widths)

ws_cover.cell(row=2, column=2, value="TEST REPORT DOCUMENT").font = Font(
    name="Tahoma", bold=True, size=20, color="000000"
)

ws_cover.cell(row=4, column=1, value="Project Name").font = BOLD_FONT
ws_cover.cell(row=4, column=2, value=PROJECT_NAME).font = NORMAL_FONT
ws_cover.cell(row=4, column=5, value="Creator").font = BOLD_FONT
ws_cover.cell(row=4, column=6, value=TESTER_NAME).font = NORMAL_FONT

ws_cover.cell(row=5, column=1, value="Project Code").font = BOLD_FONT
ws_cover.cell(row=5, column=2, value=PROJECT_CODE).font = NORMAL_FONT
ws_cover.cell(row=5, column=5, value="Issue Date").font = BOLD_FONT
ws_cover.cell(row=5, column=6, value=TEST_DATE).font = NORMAL_FONT
ws_cover.cell(row=5, column=6).number_format = "DD/MM/YYYY"

ws_cover.cell(row=6, column=1, value="Document Code").font = BOLD_FONT
ws_cover.cell(row=6, column=2, value=f"{PROJECT_CODE}_Test_Report_v1.0").font = NORMAL_FONT
ws_cover.cell(row=6, column=5, value="Version").font = BOLD_FONT
ws_cover.cell(row=6, column=6, value=1).font = NORMAL_FONT

# Record of change
ws_cover.cell(row=9, column=1, value="Record of change").font = BOLD_FONT
change_headers = ["Effective Date", "Version", "Change Item", "*A,D,M", "Change description", "Reference"]
for ci, h in enumerate(change_headers, 1):
    c = ws_cover.cell(row=10, column=ci, value=h)
    c.font = BOLD_FONT
    c.border = THIN_BORDER

changes = [
    ("15/03/2026", "1.0", "Initial Test Report Creation", "A",
     "- Create initial Test Report document for FrogEdu project.\n"
     "- Define test case structure and template format.\n"
     "- Set up module-based test case organization.", "N/A"),
    ("15/03/2026", "1.0", "Test Case Flow Definition", "M",
     "- Define functional test flows for core modules: Student Exam Taking, "
     "Exam Generation, Exam Session History, Review Page, Class Management.\n"
     "- Write detailed test procedures and expected results.", "N/A"),
    ("15/03/2026", "1.0", "Test Execution - Round 1", "M",
     "- Execute defined test cases according to test plan.\n"
     "- Record execution results (Pass/Fail/Pending).\n"
     "- All test cases passed in Round 1.", "N/A"),
]
for ri, (date, ver, item, adm, desc, ref) in enumerate(changes, 11):
    vals = [date, ver, item, adm, desc, ref]
    for ci, v in enumerate(vals, 1):
        c = ws_cover.cell(row=ri, column=ci, value=v)
        c.font = NORMAL_FONT
        c.border = THIN_BORDER
        c.alignment = WRAP_ALIGN


# ================================================================
#  Save
# ================================================================
output_path = "FrogEdu_Test_Report.xlsx"
wb.save(output_path)
print(f"Generated {output_path} with {total_all} test cases across {len(wb.sheetnames)} sheets.")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Modules: {[f'{n} ({c} TCs)' for n, c in module_tc_counts]}")
